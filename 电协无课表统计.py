import openpyxl
import numpy as np
import os

print("$$ 无课表处理\n\n")
DATAPATH = './Datas/'

files = next(os.walk(DATAPATH))[2]

# 初始化缓冲区
TargetData = [[[],[],[],[],[],[],[]],
               [[],[],[],[],[],[],[]],
               [[],[],[],[],[],[],[]],
               [[],[],[],[],[],[],[]],
               [[],[],[],[],[],[],[]],
               [[],[],[],[],[],[],[]],]

# 将数据加载到内存
def LoadData():
    totalFile = len(files)
    fileCounter = 0

    try:
        for fileName in files:
            fileCounter += 1
            wb = openpyxl.load_workbook(DATAPATH + fileName)  
            sheet = wb['Main']
    
            name = sheet.cell(row=11, column=3).value            # 读名字
            class_ = sheet.cell(row=12, column=3).value          # 读班级
    
            print(f">>> Analysis Process : {fileCounter / totalFile:>7.2%} - {name:5}  {class_}")
    
            for i in range(6):
                for j in range(7):
                    # 取出每一格数据
                    tmpVal = sheet.cell(row=4 + i, column=3 + j).value
                    # 如果是0或者没有填也就是没有课，加入缓冲区
                    if  not(tmpVal != None and tmpVal > 0):
                        TargetData[i][j].append(name)
    except Exception as e:
        print(f"\n!!!   Exception File: [ {fileName} ]\n!!!   Location:       [ {sheet.cell(row=4 + i, column=3 + j)} ]\n" )
        print("EXCEPTION: ",str(e),"\n")
        exit(0)

    return fileCounter, fileName, totalFile

# 将数据从内存中写入到磁盘
def SaveData():
    TargetFile = openpyxl.load_workbook("./TargetExcel_Template.xlsx")  
    TargetSheet = TargetFile['Main']
    
    for i in range(6):
            for j in range(7):
                print(f'>>> Saving Process : {(i*7+j)/41:>7.2%}')
                # 将缓冲区中对应表格中的数据化List为String（换行分割）载入表格
                TargetSheet.cell(row=4 + i, column=3 + j).value = "\n".join(TargetData[i][j])
    return TargetFile

# 以下内容显而易见 不需要注释了
print("$$ 开始加载数据 ##########################\n")
LoadData()
print("\n$$ 加载数据完毕 ##########################\n")

print("$$ 开始保存数据 ##########################\n")
TargetFile = SaveData()
print("\n$$ 保存数据完毕 #########################\n")

# 储存
TargetFile.save("./TargetExcel.xlsx")
print("\n$$ 汇总结果已保存到 './TargetExcel.xlsx'\n")